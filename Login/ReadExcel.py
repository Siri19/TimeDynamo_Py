import openpyxl

path="/home/siri/Documents/TimeDynamo_Automation/TD_Web_ExcelData.xlsx"
workbook=openpyxl.load_workbook(path)
sheet=workbook.get_sheet_by_name("Login_Validdata")
#sheet=workbook.active

rows=sheet.max_row
cols=sheet.max_column
print("rows-->",rows)
print("columns--->",cols)
obj=Object[rows][1]
for row in range(1,rows+1):
    for col in range(1,cols+1):
        print(sheet.cell(row=row,column=col).value,end=" ")


    print()
